import openpyxl

# Functions that count as "tricky" — Rule C
TRICKY_FUNCTIONS = [
    'VLOOKUP', 'HLOOKUP', 'INDEX', 'MATCH', 'INDIRECT',
    'OFFSET', 'SUMPRODUCT', 'ARRAYFORMULA', 'IFERROR',
    'IFNA', 'SUMIFS', 'COUNTIFS', 'AVERAGEIFS', 'CHOOSE',
    'XLOOKUP', 'FILTER', 'SORT', 'UNIQUE', 'LET', 'LAMBDA'
]

# Broken reference markers that Excel leaves in cells
BROKEN_MARKERS = ['#REF!', '#NAME?', '#VALUE!', '#DIV/0!', '#NULL!', '#N/A', '#NUM!']

# Hard cap — protects Gemini quota and Render timeout
MAX_FORMULAS = 100

def is_complex(formula):
    """Rule C: flag if over 50 chars OR contains a tricky function."""
    if len(formula) > 50:
        return True
    upper = formula.upper()
    for func in TRICKY_FUNCTIONS:
        if func in upper:
            return True
    return False

def find_broken_refs(formula):
    """Check if a formula contains any broken reference markers."""
    found = []
    upper = formula.upper()
    for marker in BROKEN_MARKERS:
        if marker in upper:
            found.append(marker)
    return found

def analyze_file(filepath):
    """Open an Excel file and return all formulas with analysis."""
    wb = openpyxl.load_workbook(filepath, data_only=False)

    results = {
        'total_formulas': 0,
        'complex_count': 0,
        'broken_count': 0,
        'formulas': []
    }

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value

                # Skip empty cells and non-formula cells
                if value is None or not isinstance(value, str) or not value.startswith('='):
                    continue

                results['total_formulas'] += 1

                # Hard cap — fail fast before we waste Gemini quota
                if results['total_formulas'] > MAX_FORMULAS:
                    wb.close()
                    raise ValueError(
                        f'File has too many formulas. Max {MAX_FORMULAS} per file. '
                        f'Try splitting your file into smaller chunks.'
                    )

                formula_info = {
                    'cell': f"{sheet_name}!{cell.coordinate}",
                    'formula': value,
                    'is_complex': is_complex(value),
                    'broken_refs': find_broken_refs(value)
                }

                if formula_info['is_complex']:
                    results['complex_count'] += 1

                if formula_info['broken_refs']:
                    results['broken_count'] += 1

                results['formulas'].append(formula_info)

    wb.close()
    return results